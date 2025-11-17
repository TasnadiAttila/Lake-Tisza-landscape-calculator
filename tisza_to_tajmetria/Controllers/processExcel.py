from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
import re
import os
import webbrowser
import plotly.graph_objects as go
from plotly.subplots import make_subplots

#TODO: patch density and nearest fix. 900% instead of 90%

FILENAME = "asd.xlsx"
OUT_FILENAME = "test_with_charts.xlsx"
METRICS = [
    "effective mesh size",
    "euclidean nearest neighbour",
    "fractal dimension index",
    "greatest patch area",
    "landscape division",
    "landscape proportion",
]
METRIC_PATTERNS = [re.compile(r"\b" + re.escape(m) + r"\b", re.IGNORECASE) for m in METRICS]


def cell_text(cell):
    v = cell.value
    if v is None:
        return ""
    return str(v).strip()


def parse_numeric(s):
    """Extract a float value from text (handles comma as decimal sep and optional percent sign).
    Returns float or None.
    """
    if s is None:
        return None
    txt = str(s).strip()
    if txt == "":
        return None
    txt = txt.replace(',', '.')
    m = re.search(r"(-?\d+(?:\.\d+)?)", txt)
    if not m:
        return None
    try:
        return float(m.group(1))
    except Exception:
        return None


def find_metric_value(sheet, r, c, max_offset=5):
    """Search to the right of (r,c) up to max_offset columns for a numeric value,
    then search the entire row, then below in the same column.
    Returns (raw_text, float_or_None).
    """
    # to the right
    for offset in range(1, max_offset + 1):
        col = c + offset
        if col > sheet.max_column:
            break
        txt = cell_text(sheet.cell(row=r, column=col))
        if not txt:
            continue
        num = parse_numeric(txt)
        if num is not None:
            return txt, num
    # anywhere in the same row
    for col in range(1, sheet.max_column + 1):
        if col == c:
            continue
        txt = cell_text(sheet.cell(row=r, column=col))
        if not txt:
            continue
        num = parse_numeric(txt)
        if num is not None:
            return txt, num
    # look below in the same column
    for row in range(r + 1, min(sheet.max_row, r + max_offset) + 1):
        txt = cell_text(sheet.cell(row=row, column=c))
        if not txt:
            continue
        num = parse_numeric(txt)
        if num is not None:
            return txt, num
    return "<no value found>", None


def find_layer_name(sheet, r, c, max_offset=5):
    """Try to find a descriptive layer name: first scan left within max_offset, then look up in column, else use sheet title."""
    # left
    for offset in range(1, max_offset + 1):
        col = c - offset
        if col < 1:
            break
        txt = cell_text(sheet.cell(row=r, column=col))
        if txt:
            return txt
    # above in same column
    for offset in range(1, max_offset + 1):
        row = r - offset
        if row < 1:
            break
        txt = cell_text(sheet.cell(row=row, column=c))
        if txt:
            return txt
    # fallback
    return sheet.title


def parse_dict_from_text(text):
    """Try to parse a dict from text like '{0.0: 74.24, 12.0: 0.13, ...}'
    Returns dict or None. Only accepts dicts with numeric keys.
    """
    if not text or '{' not in text:
        return None
    try:
        # Extract the dict part
        start = text.find('{')
        end = text.rfind('}')
        if start == -1 or end == -1:
            return None
        dict_str = text[start:end+1]
        # Use eval safely (only for dict literals)
        result = eval(dict_str, {"__builtins__": {}}, {})
        if isinstance(result, dict):
            # Validate that all keys are numeric (for land cover classes)
            for key in result.keys():
                if not isinstance(key, (int, float)):
                    # Try to convert to float
                    try:
                        float(key)
                    except (ValueError, TypeError):
                        # Not a numeric dict, skip it
                        return None
            return result
    except Exception:
        pass
    return None


def collect_metrics_from_sheet(sheet):
    """Return dict mapping layer_name -> {metric_name: float or None}
    Scans the sheet for metric labels and extracts numeric values near them.
    """
    data = {}
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            txt = cell_text(sheet.cell(row=r, column=c))
            if not txt:
                continue
            for idx, pat in enumerate(METRIC_PATTERNS):
                if pat.search(txt):
                    metric = METRICS[idx]
                    raw_val, num = find_metric_value(sheet, r, c)
                    layer = find_layer_name(sheet, r, c)
                    if layer not in data:
                        data[layer] = {m: None for m in METRICS}
                    # prefer numeric value if found
                    data[layer][metric] = num
    return data


def collect_composition_data(sheet):
    """Collect Land Cover composition data (dict format).
    Returns dict: layer_name -> {class_id: percentage}
    Excludes patch_density and nearest neighbour distance data.
    """
    composition_data = {}
    
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            txt = cell_text(sheet.cell(row=r, column=c))
            if not txt:
                continue
            
            # Skip patch_density rows
            if 'patch' in txt.lower() and 'density' in txt.lower():
                continue
            
            # Skip nearest neighbour distance rows
            if 'nearest' in txt.lower() and 'neighbour' in txt.lower():
                continue
            if 'nearest' in txt.lower() and 'neighbor' in txt.lower():  # American spelling
                continue
            
            # Look for "Raw Dict Output" or similar (but not excluded types)
            if 'raw dict' in txt.lower() or 'dict output' in txt.lower():
                # Find layer name (usually to the left)
                layer = find_layer_name(sheet, r, c)
                
                # Skip if layer name contains excluded terms
                if layer and 'nearest neighbour distance' in layer.lower():
                    continue
                if layer and 'nearest neighbor distance' in layer.lower():
                    continue
                if layer and 'patch density' in layer.lower():
                    continue
                
                # Look for dict in nearby cells (right side)
                for offset in range(1, 6):
                    col = c + offset
                    if col > sheet.max_column:
                        break
                    cell_val = cell_text(sheet.cell(row=r, column=col))
                    parsed_dict = parse_dict_from_text(cell_val)
                    if parsed_dict:
                        composition_data[layer] = parsed_dict
                        break
    
    return composition_data


def write_separate_charts(wb, sheet, data_map):
    """Create a hidden helper sheet with the Layer x Metrics table and insert one horizontal chart per metric.
    Charts are arranged in a 3x2 grid (2 columns x 3 rows), placed starting at G1 and N1, then G16/N16, G31/N31.
    Returns number of charts inserted.
    """
    if not data_map:
        return 0
    raw_name = f"_chartdata_{sheet.title}"
    data_sheet_name = re.sub(r"[\\/*?:\[\] ]", "_", raw_name)
    if data_sheet_name in wb.sheetnames:
        data_sheet = wb[data_sheet_name]
        if data_sheet.max_row > 0:
            data_sheet.delete_rows(1, data_sheet.max_row)
    else:
        data_sheet = wb.create_sheet(data_sheet_name)
    data_sheet.sheet_state = 'hidden'

    # header
    data_sheet.cell(row=1, column=1, value="Layer")
    for j, m in enumerate(METRICS, start=2):
        data_sheet.cell(row=1, column=j, value=m)

    layers = sorted(data_map.keys())
    for i, layer in enumerate(layers, start=2):
        data_sheet.cell(row=i, column=1, value=layer)
        metrics = data_map[layer]
        for j, m in enumerate(METRICS, start=2):
            val = metrics.get(m)
            if val is None:
                data_sheet.cell(row=i, column=j, value=None)
            else:
                try:
                    data_sheet.cell(row=i, column=j, value=float(val))
                except Exception:
                    data_sheet.cell(row=i, column=j, value=None)

    max_row = 1 + len(layers)
    if max_row < 2:
        return 0

    charts_created = 0
    # grid layout parameters
    col_anchors = ['G', 'N']  # left and right columns for charts
    row_spacing = 16  # increased spacing between chart rows

    for idx, m in enumerate(METRICS, start=2):
        # check if there's at least one numeric value in this metric column
        has_numeric = False
        for r in range(2, max_row + 1):
            v = data_sheet.cell(row=r, column=idx).value
            if isinstance(v, (int, float)):
                has_numeric = True
                break
        if not has_numeric:
            continue

        chart = BarChart()
        chart.type = "bar"
        chart.grouping = "clustered"
        chart.title = m
        chart.x_axis.title = "Value"
        chart.y_axis.title = "Layer"
        chart.style = 10  # choose a nicer built-in style
        chart.legend.position = 'r'

        # show values on the bars - only value, nothing else
        dLbls = DataLabelList()
        dLbls.showVal = True
        dLbls.showCatName = False
        dLbls.showSerName = False
        dLbls.showLegendKey = False
        chart.dLbls = dLbls

        # small visual tweaks
        try:
            chart.x_axis.majorGridlines = None
            chart.y_axis.majorGridlines = None
        except Exception:
            pass

        data = Reference(data_sheet, min_col=idx, min_row=1, max_col=idx, max_row=max_row)
        cats = Reference(data_sheet, min_col=1, min_row=2, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 12
        chart.height = 8

        # compute grid anchor: two columns (col_idx 0 or 1), three rows
        col_idx = charts_created % 2
        row_idx = charts_created // 2
        anchor_col = col_anchors[col_idx]
        anchor_row = 1 + (row_idx * row_spacing)
        anchor = f"{anchor_col}{anchor_row}"
        sheet.add_chart(chart, anchor)
        charts_created += 1

    # Set column M width
    sheet.column_dimensions['M'].width = 11  # adjust the value as needed (default is ~8.43)

    return charts_created


def write_composition_charts(wb, sheet, composition_data):
    """Create stacked bar charts for land cover composition data.
    composition_data is dict: layer_name -> {class_id: percentage}
    Returns number of charts created.
    """
    if not composition_data:
        return 0
    
    from openpyxl.chart import BarChart, Reference, Series
    
    # Create helper sheet for composition data
    raw_name = f"_composition_{sheet.title}"
    data_sheet_name = re.sub(r"[\\/*?:\[\] ]", "_", raw_name)
    if data_sheet_name in wb.sheetnames:
        data_sheet = wb[data_sheet_name]
        if data_sheet.max_row > 0:
            data_sheet.delete_rows(1, data_sheet.max_row)
    else:
        data_sheet = wb.create_sheet(data_sheet_name)
    data_sheet.sheet_state = 'hidden'
    
    # Collect all unique class IDs
    all_classes = set()
    for comp_dict in composition_data.values():
        for key in comp_dict.keys():
            # Only add numeric keys
            if isinstance(key, (int, float)):
                all_classes.add(key)
            else:
                try:
                    all_classes.add(float(key))
                except (ValueError, TypeError):
                    # Skip non-numeric keys
                    pass
    
    if not all_classes:
        print("No valid numeric class IDs found in composition data")
        return 0
    
    sorted_classes = sorted(all_classes)
    
    # Write header: Layer | Class_0 | Class_12 | Class_23 | ...
    data_sheet.cell(row=1, column=1, value="Layer")
    for j, class_id in enumerate(sorted_classes, start=2):
        data_sheet.cell(row=1, column=j, value=f"Class_{class_id}")
    
    # Write data
    layers = sorted(composition_data.keys())
    for i, layer in enumerate(layers, start=2):
        data_sheet.cell(row=i, column=1, value=layer)
        comp_dict = composition_data[layer]
        for j, class_id in enumerate(sorted_classes, start=2):
            # Try to get value with both float and original key
            val = comp_dict.get(class_id, comp_dict.get(str(class_id), comp_dict.get(int(class_id), 0)))
            data_sheet.cell(row=i, column=j, value=val)
    
    max_row = 1 + len(layers)
    if max_row < 2:
        return 0
    
    # Create grouped bar chart (instead of stacked)
    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "clustered"  # grouped instead of stacked
    chart.title = "Land Cover Composition"
    chart.x_axis.title = "Percentage (%)"
    chart.y_axis.title = "Layer"
    chart.style = 10
    chart.legend.position = 'r'
    
    # Add data for each class
    for idx, class_id in enumerate(sorted_classes, start=2):
        data = Reference(data_sheet, min_col=idx, min_row=1, max_col=idx, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
    
    # Set categories (layer names)
    cats = Reference(data_sheet, min_col=1, min_row=2, max_row=max_row)
    chart.set_categories(cats)
    
    chart.width = 15
    chart.height = 10
    
    # Place chart below the regular charts (around row 50)
    sheet.add_chart(chart, "B50")
    
    return 1


def create_dashboard(wb, all_data):
    """Create a dashboard sheet with all metrics visualized together.
    all_data is a dict mapping sheet_name -> layer_name -> {metric_name: float}
    """
    if not all_data:
        return 0
    
    # Create or clear dashboard sheet
    dashboard_name = "Dashboard"
    if dashboard_name in wb.sheetnames:
        dashboard = wb[dashboard_name]
        # Clear existing content
        for row in dashboard.iter_rows():
            for cell in row:
                cell.value = None
    else:
        dashboard = wb.create_sheet(dashboard_name, 0)  # insert at beginning
    
    # Create a consolidated data sheet for dashboard
    data_sheet_name = "_dashboard_data"
    if data_sheet_name in wb.sheetnames:
        data_sheet = wb[data_sheet_name]
        if data_sheet.max_row > 0:
            data_sheet.delete_rows(1, data_sheet.max_row)
    else:
        data_sheet = wb.create_sheet(data_sheet_name)
    data_sheet.sheet_state = 'hidden'
    
    # Collect all unique layers across all sheets
    all_layers = set()
    for sheet_data in all_data.values():
        all_layers.update(sheet_data.keys())
    layers = sorted(all_layers)
    
    # Write header
    data_sheet.cell(row=1, column=1, value="Layer")
    for j, m in enumerate(METRICS, start=2):
        data_sheet.cell(row=1, column=j, value=m)
    
    # Write data - aggregate from all sheets (average if multiple values)
    for i, layer in enumerate(layers, start=2):
        data_sheet.cell(row=i, column=1, value=layer)
        for j, metric in enumerate(METRICS, start=2):
            values = []
            for sheet_data in all_data.values():
                if layer in sheet_data and sheet_data[layer].get(metric) is not None:
                    values.append(sheet_data[layer][metric])
            if values:
                avg_val = sum(values) / len(values)
                data_sheet.cell(row=i, column=j, value=avg_val)
            else:
                data_sheet.cell(row=i, column=j, value=None)
    
    max_row = 1 + len(layers)
    if max_row < 2:
        return 0
    
    # Create charts in dashboard - same layout as individual sheets
    charts_created = 0
    col_anchors = ['B', 'N']
    row_spacing = 16
    
    for idx, m in enumerate(METRICS, start=2):
        # check if there's at least one numeric value
        has_numeric = False
        for r in range(2, max_row + 1):
            v = data_sheet.cell(row=r, column=idx).value
            if isinstance(v, (int, float)):
                has_numeric = True
                break
        if not has_numeric:
            continue
        
        chart = BarChart()
        chart.type = "bar"
        chart.grouping = "clustered"
        chart.title = f"{m} (Dashboard)"
        chart.x_axis.title = "Value"
        chart.y_axis.title = "Layer"
        chart.style = 10
        chart.legend.position = 'r'
        
        # show values on the bars
        dLbls = DataLabelList()
        dLbls.showVal = True
        dLbls.showCatName = False
        dLbls.showSerName = False
        dLbls.showLegendKey = False
        chart.dLbls = dLbls
        
        # visual tweaks
        try:
            chart.x_axis.majorGridlines = None
            chart.y_axis.majorGridlines = None
        except Exception:
            pass
        
        data = Reference(data_sheet, min_col=idx, min_row=1, max_col=idx, max_row=max_row)
        cats = Reference(data_sheet, min_col=1, min_row=2, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 12
        chart.height = 8
        
        # grid layout
        col_idx = charts_created % 2
        row_idx = charts_created // 2
        anchor_col = col_anchors[col_idx]
        anchor_row = 1 + (row_idx * row_spacing)
        anchor = f"{anchor_col}{anchor_row}"
        dashboard.add_chart(chart, anchor)
        charts_created += 1
    
    # Set column width for spacing
    dashboard.column_dimensions['M'].width = 2
    
    return charts_created


def create_interactive_html_dashboard(all_data, output_file, open_browser=True):
    """Create an interactive HTML dashboard using Plotly.
    all_data is a dict mapping sheet_name -> layer_name -> {metric_name: float}
    """
    if not all_data:
        print("No data to create HTML dashboard")
        return False
    
    # Collect all unique layers across all sheets
    all_layers = set()
    for sheet_data in all_data.values():
        all_layers.update(sheet_data.keys())
    layers = sorted(all_layers)
    
    # Prepare data for each metric (aggregated across sheets)
    metric_data = {metric: {} for metric in METRICS}
    
    for layer in layers:
        for metric in METRICS:
            values = []
            for sheet_data in all_data.values():
                if layer in sheet_data and sheet_data[layer].get(metric) is not None:
                    values.append(sheet_data[layer][metric])
            if values:
                avg_val = sum(values) / len(values)
                metric_data[metric][layer] = avg_val
    
    # Create subplots - 3 rows x 2 columns
    fig = make_subplots(
        rows=3, cols=2,
        subplot_titles=[m.title() for m in METRICS],
        vertical_spacing=0.12,
        horizontal_spacing=0.15
    )
    
    # Add bar charts for each metric
    positions = [(1, 1), (1, 2), (2, 1), (2, 2), (3, 1), (3, 2)]
    
    for idx, metric in enumerate(METRICS):
        if idx >= len(positions):
            break
        
        row, col = positions[idx]
        data = metric_data[metric]
        
        if not data:
            continue
        
        # Sort by value for better visualization
        sorted_items = sorted(data.items(), key=lambda x: x[1] if x[1] is not None else 0)
        layer_names = [item[0] for item in sorted_items]
        values = [item[1] for item in sorted_items]
        
        fig.add_trace(
            go.Bar(
                x=values,
                y=layer_names,
                orientation='h',
                name=metric,
                text=[f'{v:.2f}' if v is not None else 'N/A' for v in values],
                textposition='auto',
                hovertemplate='<b>%{y}</b><br>Value: %{x:.4f}<extra></extra>',
                marker=dict(
                    color=values,
                    colorscale='Viridis',
                    showscale=False
                )
            ),
            row=row, col=col
        )
        
        # Update axes
        fig.update_xaxes(title_text="Value", row=row, col=col)
        fig.update_yaxes(title_text="Layer", row=row, col=col)
    
    # Update layout
    fig.update_layout(
        title_text="<b>Landscape Metrics Dashboard</b>",
        title_font_size=24,
        title_x=0.5,
        showlegend=False,
        height=1200,
        width=1600,
        template='plotly_white',
        font=dict(size=11)
    )
    
    # Save to HTML
    try:
        fig.write_html(
            output_file,
            config={
                'displayModeBar': True,
                'displaylogo': False,
                'modeBarButtonsToRemove': ['lasso2d', 'select2d']
            }
        )
        print(f"Interactive HTML dashboard created: {output_file}")
        
        # Open in browser if requested
        if open_browser:
            webbrowser.open('file://' + os.path.abspath(output_file))
        
        return True
    except Exception as e:
        print(f"Error creating HTML dashboard: {e}")
        return False


def create_composition_html_dashboard(all_composition_data, output_file, open_browser=True):
    """Create an interactive HTML dashboard for land cover composition data using Plotly.
    all_composition_data is dict: sheet_name -> layer_name -> {class_id: percentage}
    """
    if not all_composition_data:
        print("No composition data to create HTML dashboard")
        return False
    
    # Collect all unique layers and classes
    all_layers = set()
    all_classes = set()
    for sheet_data in all_composition_data.values():
        all_layers.update(sheet_data.keys())
        for comp_dict in sheet_data.values():
            for key in comp_dict.keys():
                # Only add numeric keys
                if isinstance(key, (int, float)):
                    all_classes.add(key)
                else:
                    try:
                        all_classes.add(float(key))
                    except (ValueError, TypeError):
                        # Skip non-numeric keys
                        pass
    
    if not all_classes:
        print("No valid numeric class IDs found for HTML composition dashboard")
        return False
    
    layers = sorted(all_layers)
    sorted_classes = sorted(all_classes)
    
    # Create stacked bar chart
    fig = go.Figure()
    
    # Add a bar for each class
    for class_id in sorted_classes:
        values = []
        for layer in layers:
            # Aggregate across sheets
            total = 0
            count = 0
            for sheet_data in all_composition_data.values():
                if layer in sheet_data:
                    comp_dict = sheet_data[layer]
                    # Try to get value with different key types
                    val = comp_dict.get(class_id, comp_dict.get(str(class_id), comp_dict.get(int(class_id), None)))
                    if val is not None:
                        total += val
                        count += 1
            avg_val = total / count if count > 0 else 0
            values.append(avg_val)
        
        fig.add_trace(go.Bar(
            name=f'Class {class_id}',
            x=layers,
            y=values,
            text=[f'{v:.1f}%' for v in values],
            textposition='auto',
            hovertemplate='<b>%{x}</b><br>Class ' + str(class_id) + ': %{y:.2f}%<extra></extra>'
        ))
    
    fig.update_layout(
        title='<b>Land Cover Composition Comparison</b>',
        title_font_size=24,
        title_x=0.5,
        xaxis_title='Layer',
        yaxis_title='Percentage (%)',
        barmode='group',  # changed from 'stack' to 'group' for grouped bars
        height=700,
        width=1400,
        template='plotly_white',
        legend=dict(
            title='Land Cover Classes',
            orientation='v',
            yanchor='top',
            y=1,
            xanchor='left',
            x=1.02
        ),
        hovermode='x unified'
    )
    
    # Save to HTML
    try:
        fig.write_html(
            output_file,
            config={
                'displayModeBar': True,
                'displaylogo': False,
                'modeBarButtonsToRemove': ['lasso2d', 'select2d']
            }
        )
        print(f"Interactive composition HTML dashboard created: {output_file}")
        
        if open_browser:
            webbrowser.open('file://' + os.path.abspath(output_file))
        
        return True
    except Exception as e:
        print(f"Error creating composition HTML dashboard: {e}")
        return False


def main():
    base_dir = os.path.dirname(__file__)
    infile = os.path.join(base_dir, FILENAME)
    outfile = os.path.join(base_dir, OUT_FILENAME)
    if not os.path.exists(infile):
        print(f"Workbook not found: {infile}")
        return

    wb = load_workbook(infile)
    total_charts = 0
    all_data = {}  # collect data from all sheets for dashboard
    all_composition_data = {}  # collect composition data
    
    for sheet in wb.worksheets:
        # skip helper sheets we create and any hidden sheets
        if sheet.title.startswith("_chartdata_") or sheet.title.startswith("_composition_") or sheet.title.startswith("_dashboard") or sheet.title == "Dashboard" or getattr(sheet, 'sheet_state', '') == 'hidden':
            print(f"Skipping sheet (helper/hidden): {sheet.title}")
            continue
        
        # Collect regular metrics
        data_map = collect_metrics_from_sheet(sheet)
        if data_map:
            all_data[sheet.title] = data_map
        
        # Collect composition data
        comp_data = collect_composition_data(sheet)
        if comp_data:
            all_composition_data[sheet.title] = comp_data
            print(f"Found composition data in sheet '{sheet.title}': {len(comp_data)} layers")
        
        # Create regular charts
        charts = write_separate_charts(wb, sheet, data_map)
        if charts:
            print(f"Inserted {charts} charts into sheet '{sheet.title}'")
            total_charts += charts
        else:
            print(f"No metric data for sheet '{sheet.title}'")
        
        # Create composition charts
        comp_charts = write_composition_charts(wb, sheet, comp_data)
        if comp_charts:
            print(f"Inserted {comp_charts} composition chart(s) into sheet '{sheet.title}'")
            total_charts += comp_charts
    
    # Create dashboard with all collected data
    if all_data:
        dashboard_charts = create_dashboard(wb, all_data)
        if dashboard_charts:
            print(f"Created Dashboard with {dashboard_charts} charts")
            total_charts += dashboard_charts
        
        # Create interactive HTML dashboard
        html_output = os.path.join(base_dir, "dashboard_interactive.html")
        create_interactive_html_dashboard(all_data, html_output)
    
    # Create composition HTML dashboard
    if all_composition_data:
        comp_html_output = os.path.join(base_dir, "composition_dashboard.html")
        create_composition_html_dashboard(all_composition_data, comp_html_output)
    
    if total_charts > 0:
        try:
            wb.save(outfile)
            print(f"Saved workbook with {total_charts} new charts: {outfile}")
        except PermissionError:
            print(f"Permission denied when saving to {outfile}. Is the file open? Try closing it or choose a different name.")
    else:
        print("No changes to save.")


if __name__ == '__main__':
    main()
